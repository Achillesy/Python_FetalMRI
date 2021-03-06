from openpyxl import load_workbook

file = '../FetalData/20210301/Accession number lookup.xlsx'
sheetname = 'Sheet1'

def read_excel(filename, sheetname):
    wb = load_workbook(filename)
    ws = wb[sheetname]
    return ws.iter_rows()
 
rows = read_excel(file, sheetname)
for row in rows:
    PseudoAcc = ''.join([c for c in str(row[0].value) if c in '1234567890'])
    OrigAcc = ''.join([c for c in str(row[1].value) if c in '1234567890'])
    PseudoName = str(row[2].value)
    print(PseudoAcc+' '+OrigAcc+' '+PseudoName)
