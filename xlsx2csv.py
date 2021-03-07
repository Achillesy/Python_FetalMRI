#!/usr/bin/env python

#
# Instructions:
# Convert xlsx file to csv
#
# \author     Xuchu Liu (xuchu_liu@rush.edu)
# \date       03/06/2021
#

from openpyxl import load_workbook
import csv

xlsxFile = '../FetalData/20210301/Accession number lookup.xlsx'
sheetname = 'Sheet1'
csvFile = '../FetalData/20210301/Accession number lookup.csv'

def read_excel(filename, sheetname):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheetname]
    return ws.iter_rows()

cfile = open(csvFile, 'w', newline='', encoding='utf-8-sig')
writer = csv.writer(cfile)
rows = read_excel(xlsxFile, sheetname)
for row in rows:
    PseudoAcc = ''.join([c for c in str(row[0].value) if c in '1234567890'])
    if len(PseudoAcc) < 1 :
        continue
    OrigAcc = ''.join([c for c in str(row[1].value) if c in '1234567890'])
    PseudoName = str(row[2].value)
    print([PseudoAcc, OrigAcc, PseudoName]) # Debug
    writer.writerow([PseudoAcc, OrigAcc, PseudoName])
cfile.close()

# Excel UUID Function
# =LOWER(CONCATENATE(DEC2HEX(RANDBETWEEN(0,POWER(16,8)),8),"-",DEC2HEX(RANDBETWEEN(0,POWER(16,4)),4),"-","4",DEC2HEX(RANDBETWEEN(0,POWER(16,3)),3),"-",DEC2HEX(RANDBETWEEN(8,11)),DEC2HEX(RANDBETWEEN(0,POWER(16,3)),3),"-",DEC2HEX(RANDBETWEEN(0,POWER(16,8)),8),DEC2HEX(RANDBETWEEN(0,POWER(16,4)),4)))
