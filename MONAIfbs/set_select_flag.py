#!/usr/bin/env python
#
# Instructions:
# Set series table State for be selected. 
#
# \author     Xuchu Liu (xuchu_liu@rush.edu)
# \date       03/11/2021
#

import sys
import fetaldb
from openpyxl import load_workbook

if len(sys.argv) != 6:
    print("""
Set series table State for be selected. 

Usage: %s seria_list.xlsx sheetname PseudoAcc SeriesNum State
""" % (sys.argv[0]))
    exit(-1)

xlsxname = sys.argv[1]
sheetname = sys.argv[2]
PseudoAccIdx = int(sys.argv[3])
SeriesNumIdx = int(sys.argv[4])
StateIdx = int(sys.argv[5])

def read_excel(filename, sheetname):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheetname]
    return ws.iter_rows()

dbFile = "FetalMRIsqlite3.db"
conn = fetaldb.FetalDB(dbFile)

rows = read_excel(xlsxname, sheetname)
for row in rows:
    PseudoAcc = row[PseudoAccIdx].value
    SeriesNum = row[SeriesNumIdx].value
    State = row[StateIdx].value
    conn.update_series_state_recon(str(PseudoAcc).zfill(6), str(SeriesNum), str(State))

conn.close()
